import os
import re

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font

from util.handle_file import Handlefile

list_pwd = os.listdir(os.path.join(os.getcwd(), 'original_data'))
date = re.findall(r"_(\d+).", list_pwd[1])[0]
header_list_rate = ['审批类型', '总次数', '通过次数', '拒绝次数', '审批中', '已撤回', '通过率（通过/（通过+拒绝）']
header_list_no = ['申请编号', '申请类型']
rate_file_name = os.path.join(os.path.abspath("result/"), f'审批通过率数据文档_{date}.xlsx')


def deal_data(file_name):
    hd = Handlefile()
    file_path = os.path.join(os.getcwd(), f'original_data/{file_name}')
    sheet = 'Sheet1'
    data_list = hd.read_execl(file_path, sheet)
    list_result = os.listdir(os.path.join(os.getcwd(), 'result'))
    if f'审批通过率数据文档_{date}.xlsx' not in list_result:
        init_excel()
    wb = load_workbook(rate_file_name)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]
    ws.append(rate(data_list))
    ws2 = wb[sheets[1]]
    for item in failure_no('已拒绝', data_list):
        ws2.append(item)
    wb.save(rate_file_name)


def init_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '审批情况汇总'
    ws.row_dimensions[1].height = 20
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["G"].width = 30
    for i in range(len(header_list_rate)):
        ws.cell(1, i + 1).value = header_list_rate[i]
        ws.cell(1, i + 1).font = Font(bold=True, size=12)
    ws2 = wb.create_sheet('拒绝编号')
    ws2.row_dimensions[1].height = 20
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20
    for i in range(len(header_list_no)):
        ws2.cell(1, i + 1).value = header_list_no[i]
        ws2.cell(1, i + 1).font = Font(bold=True, size=12)
    wb.save(rate_file_name)


def rate(data_list):
    result_list = list()
    result_list.append(data_list[1][1])
    result_list.append(len(data_list) - 1)
    result_list.append(key_times('已同意', 2, data_list))
    result_list.append(key_times('已拒绝', 2, data_list))
    result_list.append(key_times('审批中', 2, data_list))
    result_list.append(key_times('已撤回', 2, data_list))
    pass_rate = round(
        key_times('已同意', 2, data_list) / (key_times('已同意', 2, data_list) + key_times('已拒绝', 2, data_list)), 2)
    result_list.append(pass_rate)

    return result_list


def key_times(key, cow, data_list):
    times = 0
    for i in range(len(data_list) - 1):
        if data_list[i + 1][cow] == key:
            times += 1
    return times


def failure_no(key, data_list):
    result_list = list()
    for i in range(len(data_list) - 1):
        if data_list[i + 1][2] == key:
            result_list.append([data_list[i + 1][0], data_list[i + 1][1]])
    return result_list


for excel_file in list_pwd:
    if '内部提测' in excel_file:
        deal_data(f'内部提测_全部_{date}.xlsx')
    elif '正式环境发版' in excel_file:
        deal_data(f'正式环境发版_全部_{date}.xlsx')
    elif '研发计划变更' in excel_file:
        deal_data(f'研发计划变更_全部_{date}.xlsx')
    elif '集成提测' in excel_file:
        deal_data(f'集成提测_全部_{date}.xlsx')
    elif '预发布环境发版' in excel_file:
        deal_data(f'预发布环境发版_全部_{date}.xlsx')
    else:
        print('无法识别文件:\t' + excel_file)
